"""
Report generation for financial closeout reconciliation.

Generates Excel workbooks and PDF summaries from reconciliation data.
"""

import logging
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Optional

from .models import (
    CloseoutItem,
    CloseoutReport,
    CostCodeMapping,
    ItemStatus,
    ReconciliationResult,
    Severity,
)

logger = logging.getLogger(__name__)


# Color constants for Excel
COLORS = {
    "green": "92D050",  # Reconciled/OK
    "yellow": "FFFF00",  # Warning
    "red": "FF6B6B",  # Critical
    "light_blue": "B8CCE4",  # Header
    "light_gray": "F2F2F2",  # Alternate row
    "white": "FFFFFF",
}


def severity_to_color(severity: Severity) -> str:
    """Map severity to color hex code."""
    color_map = {
        Severity.INFO: COLORS["green"],
        Severity.WARNING: COLORS["yellow"],
        Severity.CRITICAL: COLORS["red"],
    }
    return color_map.get(severity, COLORS["white"])


def format_currency(value: Optional[Decimal]) -> str:
    """Format decimal as currency string."""
    if value is None:
        return "-"
    return f"${value:,.2f}"


def format_percentage(value: float) -> str:
    """Format float as percentage string."""
    return f"{value:.1f}%"


class ExcelReportGenerator:
    """
    Generate Excel workbook from closeout report.

    Creates a multi-sheet workbook with reconciliation data,
    color-coded by severity with conditional formatting.
    """

    def __init__(self, report: CloseoutReport):
        """
        Initialize generator with report data.

        Args:
            report: CloseoutReport to generate from
        """
        self.report = report
        self._wb = None
        self._styles = {}

    def _init_workbook(self):
        """Initialize openpyxl workbook and styles."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel report generation")

        self._wb = Workbook()
        self._Alignment = Alignment
        self._Border = Border
        self._Font = Font
        self._PatternFill = PatternFill
        self._Side = Side
        self._get_column_letter = get_column_letter

        # Define styles
        thin_border = Side(style="thin", color="000000")
        self._styles = {
            "header": {
                "font": Font(bold=True, color="FFFFFF"),
                "fill": PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                ),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": Border(
                    left=thin_border,
                    right=thin_border,
                    top=thin_border,
                    bottom=thin_border,
                ),
            },
            "currency": {
                "alignment": Alignment(horizontal="right"),
                "number_format": "$#,##0.00",
            },
            "percentage": {
                "alignment": Alignment(horizontal="right"),
                "number_format": "0.0%",
            },
            "green": {
                "fill": PatternFill(
                    start_color=COLORS["green"],
                    end_color=COLORS["green"],
                    fill_type="solid",
                )
            },
            "yellow": {
                "fill": PatternFill(
                    start_color=COLORS["yellow"],
                    end_color=COLORS["yellow"],
                    fill_type="solid",
                )
            },
            "red": {
                "fill": PatternFill(
                    start_color=COLORS["red"], end_color=COLORS["red"], fill_type="solid"
                )
            },
        }

    def _apply_header_style(self, cell):
        """Apply header style to a cell."""
        for attr, value in self._styles["header"].items():
            setattr(cell, attr, value)

    def _apply_severity_color(self, cell, severity: Severity):
        """Apply color based on severity."""
        color_map = {
            Severity.INFO: "green",
            Severity.WARNING: "yellow",
            Severity.CRITICAL: "red",
        }
        style_name = color_map.get(severity, "green")
        if style_name in self._styles:
            cell.fill = self._styles[style_name]["fill"]

    def _auto_column_width(self, ws):
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _create_summary_sheet(self):
        """Create executive summary sheet."""
        ws = self._wb.active
        ws.title = "Executive Summary"

        summary = self.report.project_summary

        # Title
        ws["A1"] = "FINANCIAL CLOSEOUT SUMMARY"
        ws["A1"].font = self._Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        # Project info
        ws["A3"] = "Project:"
        ws["B3"] = summary.project_name
        ws["A4"] = "Project Number:"
        ws["B4"] = summary.project_number or "-"
        ws["A5"] = "Generated:"
        ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M")

        # Key metrics section
        ws["A7"] = "KEY FINANCIAL METRICS"
        ws["A7"].font = self._Font(bold=True, size=12)
        ws.merge_cells("A7:D7")

        metrics = [
            ("Revised Contract Value", summary.revised_contract_value),
            ("Total Committed", summary.total_committed),
            ("Billed by Subcontractors", summary.total_billed_by_subs),
            ("Paid to Subcontractors", summary.total_paid_to_subs),
            ("Retention Held", summary.sub_retention_held),
        ]

        row = 9
        for label, value in metrics:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = float(value)
            ws[f"B{row}"].number_format = "$#,##0.00"
            row += 1

        # Reconciliation status
        row += 1
        ws[f"A{row}"] = "RECONCILIATION STATUS"
        ws[f"A{row}"].font = self._Font(bold=True, size=12)
        row += 2

        status_items = [
            ("Items Reconciled", summary.reconciled_items, "green"),
            ("Warnings", summary.warning_items, "yellow"),
            ("Critical Issues", summary.critical_items, "red"),
            ("Open Closeout Items", summary.open_closeout_items, None),
        ]

        for label, value, color in status_items:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            if color:
                ws[f"B{row}"].fill = self._styles[color]["fill"]
            row += 1

        row += 1
        ws[f"A{row}"] = "Estimated Exposure:"
        ws[f"A{row}"].font = self._Font(bold=True)
        ws[f"B{row}"] = float(summary.estimated_exposure)
        ws[f"B{row}"].number_format = "$#,##0.00"
        ws[f"B{row}"].fill = self._styles["red"]["fill"] if summary.estimated_exposure > 0 else self._styles["green"]["fill"]

        # Executive summary text
        if self.report.executive_summary:
            row += 3
            ws[f"A{row}"] = "EXECUTIVE SUMMARY"
            ws[f"A{row}"].font = self._Font(bold=True, size=12)
            row += 2
            ws[f"A{row}"] = self.report.executive_summary
            ws.merge_cells(f"A{row}:F{row + 10}")
            ws[f"A{row}"].alignment = self._Alignment(wrap_text=True, vertical="top")

        self._auto_column_width(ws)

    def _create_reconciliation_sheet(
        self,
        sheet_name: str,
        results: list[ReconciliationResult],
    ):
        """Create a reconciliation data sheet."""
        ws = self._wb.create_sheet(sheet_name)

        # Headers
        headers = [
            "ID",
            "Description",
            "Vendor",
            "Procore Value",
            "QB Value",
            "Variance",
            "Variance %",
            "Severity",
            "Notes",
            "Action Required",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Data rows
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=result.id)
            ws.cell(row=row_idx, column=2, value=result.item_description)
            ws.cell(row=row_idx, column=3, value=result.vendor or "-")

            # Currency columns
            procore_cell = ws.cell(
                row=row_idx,
                column=4,
                value=float(result.procore_value) if result.procore_value else None,
            )
            procore_cell.number_format = "$#,##0.00"

            qb_cell = ws.cell(
                row=row_idx,
                column=5,
                value=float(result.qb_value) if result.qb_value else None,
            )
            qb_cell.number_format = "$#,##0.00"

            variance_cell = ws.cell(
                row=row_idx, column=6, value=float(result.variance)
            )
            variance_cell.number_format = "$#,##0.00"
            self._apply_severity_color(variance_cell, result.severity)

            ws.cell(row=row_idx, column=7, value=result.variance_pct / 100)
            ws.cell(row=row_idx, column=7).number_format = "0.0%"

            severity_cell = ws.cell(
                row=row_idx, column=8, value=result.severity.value.upper()
            )
            self._apply_severity_color(severity_cell, result.severity)

            ws.cell(row=row_idx, column=9, value=result.notes)
            ws.cell(
                row=row_idx, column=10, value="Yes" if result.requires_action else ""
            )

        self._auto_column_width(ws)

    def _create_closeout_items_sheet(self):
        """Create closeout items sheet."""
        ws = self._wb.create_sheet("Open Items")

        headers = [
            "ID",
            "Category",
            "Description",
            "Vendor",
            "Amount at Risk",
            "Action Required",
            "Responsible Party",
            "Status",
            "Priority",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        ws.freeze_panes = "A2"

        for row_idx, item in enumerate(self.report.closeout_items, 2):
            ws.cell(row=row_idx, column=1, value=item.id)
            ws.cell(row=row_idx, column=2, value=item.category)
            ws.cell(row=row_idx, column=3, value=item.description)
            ws.cell(row=row_idx, column=4, value=item.vendor or "-")

            amount_cell = ws.cell(
                row=row_idx, column=5, value=float(item.amount_at_risk)
            )
            amount_cell.number_format = "$#,##0.00"

            ws.cell(row=row_idx, column=6, value=item.action_required)
            ws.cell(row=row_idx, column=7, value=item.responsible_party or "-")
            ws.cell(row=row_idx, column=8, value=item.status.value)
            ws.cell(row=row_idx, column=9, value=item.priority)

            # Color by priority
            if item.priority == 1:
                for col in range(1, 10):
                    ws.cell(row=row_idx, column=col).fill = self._styles["red"]["fill"]
            elif item.priority == 2:
                for col in range(1, 10):
                    ws.cell(row=row_idx, column=col).fill = self._styles["yellow"]["fill"]

        self._auto_column_width(ws)

    def _create_mapping_sheet(self):
        """Create cost code mapping sheet."""
        if not self.report.cost_code_mappings:
            return

        ws = self._wb.create_sheet("Cost Code Mapping")

        headers = [
            "Procore Code",
            "Description",
            "CSI Division",
            "QB Account ID",
            "QB Account Name",
            "Confidence",
            "Verified",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        ws.freeze_panes = "A2"

        for row_idx, mapping in enumerate(self.report.cost_code_mappings, 2):
            ws.cell(row=row_idx, column=1, value=mapping.procore_cost_code)
            ws.cell(row=row_idx, column=2, value=mapping.procore_description)
            ws.cell(row=row_idx, column=3, value=mapping.csi_division)
            ws.cell(row=row_idx, column=4, value=mapping.qb_account_id or "-")
            ws.cell(row=row_idx, column=5, value=mapping.qb_account_name or "-")
            ws.cell(row=row_idx, column=6, value=mapping.confidence)
            ws.cell(row=row_idx, column=6).number_format = "0%"
            ws.cell(row=row_idx, column=7, value="Yes" if mapping.manually_verified else "")

        self._auto_column_width(ws)

    def generate(self, output_path: Path) -> Path:
        """
        Generate Excel workbook and save to file.

        Args:
            output_path: Path to save the workbook

        Returns:
            Path to the generated file
        """
        self._init_workbook()

        # Create sheets
        self._create_summary_sheet()

        if self.report.commitment_reconciliation:
            self._create_reconciliation_sheet(
                "Commitment Reconciliation", self.report.commitment_reconciliation
            )

        if self.report.invoice_reconciliation:
            self._create_reconciliation_sheet(
                "Invoice Detail", self.report.invoice_reconciliation
            )

        if self.report.change_order_reconciliation:
            self._create_reconciliation_sheet(
                "Change Order Log", self.report.change_order_reconciliation
            )

        if self.report.retention_reconciliation:
            self._create_reconciliation_sheet(
                "Retention Summary", self.report.retention_reconciliation
            )

        if self.report.budget_reconciliation:
            self._create_reconciliation_sheet(
                "Budget vs Actual", self.report.budget_reconciliation
            )

        if self.report.closeout_items:
            self._create_closeout_items_sheet()

        if self.report.cost_code_mappings:
            self._create_mapping_sheet()

        # Save
        output_path = Path(output_path)
        if not output_path.suffix:
            output_path = output_path.with_suffix(".xlsx")

        self._wb.save(output_path)
        logger.info(f"Generated Excel report: {output_path}")

        return output_path


class PDFReportGenerator:
    """
    Generate PDF summary report from closeout data.

    Creates a polished 1-2 page PDF with key metrics and action items.
    """

    def __init__(self, report: CloseoutReport):
        """
        Initialize generator with report data.

        Args:
            report: CloseoutReport to generate from
        """
        self.report = report

    def generate(self, output_path: Path) -> Path:
        """
        Generate PDF report and save to file.

        Args:
            output_path: Path to save the PDF

        Returns:
            Path to the generated file
        """
        try:
            return self._generate_with_reportlab(output_path)
        except ImportError:
            logger.warning("reportlab not available, trying weasyprint")
            try:
                return self._generate_with_weasyprint(output_path)
            except ImportError:
                logger.error("Neither reportlab nor weasyprint available")
                raise ImportError(
                    "PDF generation requires either reportlab or weasyprint"
                )

    def _generate_with_reportlab(self, output_path: Path) -> Path:
        """Generate PDF using reportlab."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        output_path = Path(output_path)
        if not output_path.suffix:
            output_path = output_path.with_suffix(".pdf")

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=letter,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        styles = getSampleStyleSheet()
        elements = []

        summary = self.report.project_summary

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=12,
        )
        elements.append(Paragraph("Financial Closeout Summary", title_style))

        # Project info
        project_info = f"""
        <b>Project:</b> {summary.project_name}<br/>
        <b>Project Number:</b> {summary.project_number or '-'}<br/>
        <b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}
        """
        elements.append(Paragraph(project_info, styles["Normal"]))
        elements.append(Spacer(1, 0.25 * inch))

        # Key metrics table
        elements.append(Paragraph("<b>Key Financial Metrics</b>", styles["Heading2"]))

        metrics_data = [
            ["Metric", "Value"],
            ["Revised Contract Value", format_currency(summary.revised_contract_value)],
            ["Total Committed", format_currency(summary.total_committed)],
            ["Billed by Subcontractors", format_currency(summary.total_billed_by_subs)],
            ["Paid to Subcontractors", format_currency(summary.total_paid_to_subs)],
            ["Retention Held", format_currency(summary.sub_retention_held)],
        ]

        metrics_table = Table(metrics_data, colWidths=[3 * inch, 2 * inch])
        metrics_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.gray),
                ]
            )
        )
        elements.append(metrics_table)
        elements.append(Spacer(1, 0.25 * inch))

        # Reconciliation status
        elements.append(Paragraph("<b>Reconciliation Status</b>", styles["Heading2"]))

        status_data = [
            ["Status", "Count"],
            ["Items Reconciled", str(summary.reconciled_items)],
            ["Warnings", str(summary.warning_items)],
            ["Critical Issues", str(summary.critical_items)],
            ["Open Closeout Items", str(summary.open_closeout_items)],
            ["Estimated Exposure", format_currency(summary.estimated_exposure)],
        ]

        status_table = Table(status_data, colWidths=[3 * inch, 2 * inch])
        status_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.gray),
                    # Color-code rows
                    ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#FFFF99")),  # Warning
                    ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#FF9999")),  # Critical
                ]
            )
        )
        elements.append(status_table)
        elements.append(Spacer(1, 0.25 * inch))

        # Top action items
        if self.report.closeout_items:
            elements.append(
                Paragraph("<b>Top Action Items</b>", styles["Heading2"])
            )

            items_data = [["Priority", "Description", "Amount", "Action"]]
            for item in self.report.closeout_items[:10]:
                items_data.append(
                    [
                        str(item.priority),
                        item.description[:40] + "..." if len(item.description) > 40 else item.description,
                        format_currency(item.amount_at_risk),
                        item.action_required[:30] + "..." if len(item.action_required) > 30 else item.action_required,
                    ]
                )

            items_table = Table(
                items_data, colWidths=[0.6 * inch, 2.5 * inch, 1.2 * inch, 2.2 * inch]
            )
            items_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.gray),
                    ]
                )
            )
            elements.append(items_table)
            elements.append(Spacer(1, 0.25 * inch))

        # Executive summary
        if self.report.executive_summary:
            elements.append(Paragraph("<b>Executive Summary</b>", styles["Heading2"]))
            # Wrap long text
            exec_style = ParagraphStyle(
                "ExecSummary",
                parent=styles["Normal"],
                fontSize=9,
                leading=12,
            )
            # Replace newlines with <br/> for proper rendering
            summary_text = self.report.executive_summary.replace("\n", "<br/>")
            elements.append(Paragraph(summary_text, exec_style))

        # Build PDF
        doc.build(elements)
        logger.info(f"Generated PDF report: {output_path}")

        return output_path

    def _generate_with_weasyprint(self, output_path: Path) -> Path:
        """Generate PDF using weasyprint (HTML to PDF)."""
        from weasyprint import HTML

        summary = self.report.project_summary

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; font-size: 11pt; }}
                h1 {{ color: #4472C4; border-bottom: 2px solid #4472C4; }}
                h2 {{ color: #4472C4; margin-top: 20px; }}
                table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
                th {{ background-color: #4472C4; color: white; padding: 8px; text-align: left; }}
                td {{ padding: 6px 8px; border: 1px solid #ddd; }}
                .currency {{ text-align: right; }}
                .warning {{ background-color: #FFFF99; }}
                .critical {{ background-color: #FF9999; }}
                .success {{ background-color: #92D050; }}
            </style>
        </head>
        <body>
            <h1>Financial Closeout Summary</h1>

            <p>
                <strong>Project:</strong> {summary.project_name}<br>
                <strong>Project Number:</strong> {summary.project_number or '-'}<br>
                <strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M')}
            </p>

            <h2>Key Financial Metrics</h2>
            <table>
                <tr><th>Metric</th><th class="currency">Value</th></tr>
                <tr><td>Revised Contract Value</td><td class="currency">{format_currency(summary.revised_contract_value)}</td></tr>
                <tr><td>Total Committed</td><td class="currency">{format_currency(summary.total_committed)}</td></tr>
                <tr><td>Billed by Subcontractors</td><td class="currency">{format_currency(summary.total_billed_by_subs)}</td></tr>
                <tr><td>Paid to Subcontractors</td><td class="currency">{format_currency(summary.total_paid_to_subs)}</td></tr>
                <tr><td>Retention Held</td><td class="currency">{format_currency(summary.sub_retention_held)}</td></tr>
            </table>

            <h2>Reconciliation Status</h2>
            <table>
                <tr><th>Status</th><th>Count</th></tr>
                <tr class="success"><td>Items Reconciled</td><td>{summary.reconciled_items}</td></tr>
                <tr class="warning"><td>Warnings</td><td>{summary.warning_items}</td></tr>
                <tr class="critical"><td>Critical Issues</td><td>{summary.critical_items}</td></tr>
                <tr><td>Open Closeout Items</td><td>{summary.open_closeout_items}</td></tr>
                <tr><td><strong>Estimated Exposure</strong></td><td class="currency"><strong>{format_currency(summary.estimated_exposure)}</strong></td></tr>
            </table>
        """

        # Add top action items
        if self.report.closeout_items:
            html_content += """
            <h2>Top Action Items</h2>
            <table>
                <tr><th>Priority</th><th>Description</th><th class="currency">Amount</th><th>Action</th></tr>
            """
            for item in self.report.closeout_items[:10]:
                row_class = "critical" if item.priority == 1 else ("warning" if item.priority == 2 else "")
                html_content += f"""
                <tr class="{row_class}">
                    <td>{item.priority}</td>
                    <td>{item.description[:50]}</td>
                    <td class="currency">{format_currency(item.amount_at_risk)}</td>
                    <td>{item.action_required[:40]}</td>
                </tr>
                """
            html_content += "</table>"

        # Add executive summary
        if self.report.executive_summary:
            html_content += f"""
            <h2>Executive Summary</h2>
            <p>{self.report.executive_summary.replace(chr(10), '<br>')}</p>
            """

        html_content += """
        </body>
        </html>
        """

        output_path = Path(output_path)
        if not output_path.suffix:
            output_path = output_path.with_suffix(".pdf")

        HTML(string=html_content).write_pdf(output_path)
        logger.info(f"Generated PDF report: {output_path}")

        return output_path


def generate_reports(
    report: CloseoutReport,
    output_dir: Path,
    project_name: Optional[str] = None,
    generate_excel: bool = True,
    generate_pdf: bool = True,
) -> dict[str, Path]:
    """
    Generate all report formats.

    Args:
        report: CloseoutReport data
        output_dir: Directory to save reports
        project_name: Project name for filename
        generate_excel: Whether to generate Excel report
        generate_pdf: Whether to generate PDF report

    Returns:
        Dictionary mapping format to file path
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    name_part = project_name or report.project_summary.project_name
    name_part = "".join(c for c in name_part if c.isalnum() or c in " -_")[:50]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"closeout_{name_part}_{timestamp}"

    results = {}

    if generate_excel:
        try:
            excel_path = output_dir / f"{base_name}.xlsx"
            generator = ExcelReportGenerator(report)
            results["excel"] = generator.generate(excel_path)
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")

    if generate_pdf:
        try:
            pdf_path = output_dir / f"{base_name}.pdf"
            generator = PDFReportGenerator(report)
            results["pdf"] = generator.generate(pdf_path)
        except Exception as e:
            logger.error(f"Failed to generate PDF report: {e}")

    return results
